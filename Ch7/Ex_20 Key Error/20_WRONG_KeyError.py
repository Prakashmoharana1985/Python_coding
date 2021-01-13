# 20_WRONG_KeyError.py
from openpyxl import load_workbook, styles
wb = load_workbook('before.xlsx', data_only=True)
ws = wb["Sheet1"]

ft = styles.Font(color='4F81BD', bold=True)
ws['A1'].font = ft
ws.cell(row=1, column=1).value = 'Heading 1'
ws.column_dimensions['A'].width = 12

for row in ws.iter_rows():
    for cell in row:
        print("Looping through data")
