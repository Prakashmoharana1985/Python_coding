# Ex_6_CORRECT.py
from openpyxl import load_workbook
wb1 = load_workbook('Before.xlsx', data_only=True)
ws1 = wb1["ExportedData"]

bfr = 2
while bfr <= ws1.max_row:
    bfritem = ws1.cell(row=bfr, column=2).value
    print(bfritem)
    bfr = bfr + 1
